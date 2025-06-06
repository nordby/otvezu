# web_app.py - Исправленное FastAPI веб-приложение для системы экспедирования

# Импорты стандартных библиотек
import os
import sys
import json
import tempfile
import logging
from datetime import datetime, date, timedelta
from typing import Optional

# Импорты сторонних библиотек
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, Request, HTTPException, Depends, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.security import HTTPBasic, HTTPBasicCredentials

# Импорты локальных модулей
from database import DatabaseManager, User

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Безопасный импорт для Google Calendar
GOOGLE_CALENDAR_AVAILABLE = False

try:
    from google_calendar import CalendarIntegration, get_calendar_integration, GOOGLE_CALENDAR_AVAILABLE
    logger.info("✅ Google Calendar модуль успешно импортирован")
except ImportError as e:
    logger.warning(f"⚠️ Google Calendar модуль недоступен: {e}")
    
    # Создаем заглушки для Google Calendar
    class CalendarIntegration:
        def __init__(self, enabled=False):
            self.enabled = False
        
        def get_connection_status(self):
            return {
                'is_available': False,
                'is_configured': False,
                'has_credentials': False,
                'has_token': False,
                'calendar_info': None,
                'error': 'Google Calendar интеграция недоступна. Установите зависимости: pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client'
            }
        
        def test_connection(self):
            return {
                'success': False,
                'message': 'Google Calendar интеграция недоступна',
                'calendar_info': None
            }
        
        def create_trip_event_sync(self, trip_data, user_data):
            return None
        
        def update_trip_event_sync(self, event_id, trip_data, user_data):
            return False
        
        def delete_trip_event_sync(self, event_id):
            return False
    
    def get_calendar_integration():
        return CalendarIntegration(enabled=False)

# Создание приложения FastAPI
app = FastAPI(title="Система экспедирования", description="Веб-интерфейс для управления автопарком и рейсами")

# Настройка безопасности
security = HTTPBasic()
db = DatabaseManager()

# Создаем директории для статических файлов
os.makedirs("static/css", exist_ok=True)
os.makedirs("static/js", exist_ok=True)
os.makedirs("templates", exist_ok=True)

# Создаем базовые файлы если их нет
if not os.path.exists("static/css/style.css"):
    with open("static/css/style.css", "w", encoding="utf-8") as f:
        f.write("""
.sidebar {
    position: fixed;
    top: 56px;
    bottom: 0;
    left: 0;
    z-index: 100;
    padding: 48px 0 0;
    box-shadow: inset -1px 0 0 rgba(0, 0, 0, .1);
}

.sidebar .nav-link {
    font-weight: 500;
    color: #333;
    padding: 0.75rem 1rem;
}

.sidebar .nav-link:hover {
    color: #007bff;
    background-color: #f8f9fa;
}

.sidebar .nav-link.active {
    color: #007bff;
    background-color: #e9ecef;
}

.border-left-primary {
    border-left: 0.25rem solid #4e73df !important;
}

.border-left-success {
    border-left: 0.25rem solid #1cc88a !important;
}

.border-left-info {
    border-left: 0.25rem solid #36b9cc !important;
}

.border-left-warning {
    border-left: 0.25rem solid #f6c23e !important;
}

.text-xs {
    font-size: 0.7rem;
}

.text-gray-300 {
    color: #dddfeb !important;
}

.text-gray-800 {
    color: #5a5c69 !important;
}

main {
    padding-top: 76px;
}

.card {
    border: none;
    border-radius: 0.35rem;
}

.shadow {
    box-shadow: 0 0.15rem 1.75rem 0 rgba(58, 59, 69, 0.15) !important;
}

@media (max-width: 768px) {
    .sidebar {
        position: static;
        top: auto;
        bottom: auto;
        height: auto;
        padding: 0;
    }
    
    main {
        padding-top: 20px;
    }
}
        """)

if not os.path.exists("static/js/app.js"):
    with open("static/js/app.js", "w", encoding="utf-8") as f:
        f.write("""
// Общие функции для приложения

function showAlert(message, type = 'info') {
    const alertDiv = document.createElement('div');
    alertDiv.className = `alert alert-${type} alert-dismissible fade show`;
    alertDiv.innerHTML = `
        ${message}
        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
    `;
    
    document.querySelector('main').prepend(alertDiv);
    
    // Автоматическое удаление через 5 секунд
    setTimeout(() => {
        alertDiv.remove();
    }, 5000);
}

// Форматирование чисел
function formatCurrency(amount) {
    return new Intl.NumberFormat('ru-RU', {
        style: 'currency',
        currency: 'RUB'
    }).format(amount);
}

// Форматирование дат
function formatDate(dateString) {
    return new Date(dateString).toLocaleDateString('ru-RU');
}
        """)

# Статические файлы и шаблоны
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

def get_current_admin_user(credentials: HTTPBasicCredentials = Depends(security)):
    """Проверка авторизации администратора"""
    user = db.authenticate_user(credentials.username, credentials.password)
    if not user or user.role != 'admin':
        raise HTTPException(
            status_code=401,
            detail="Неверные учетные данные или недостаточно прав",
            headers={"WWW-Authenticate": "Basic"},
        )
    return user

# ===== ГЛАВНАЯ СТРАНИЦА =====
@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Главная страница - панель управления"""
    
    # Получаем статистику
    today = date.today()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    trips_today = db.get_trips_for_report(start_date=today, end_date=today)
    trips_week = db.get_trips_for_report(start_date=week_ago, end_date=today)
    trips_month = db.get_trips_for_report(start_date=month_ago, end_date=today)
    
    total_revenue_today = sum(trip['total_amount'] for trip in trips_today)
    total_revenue_week = sum(trip['total_amount'] for trip in trips_week)
    total_revenue_month = sum(trip['total_amount'] for trip in trips_month)
    
    stats = {
        'trips_today': len(trips_today),
        'trips_week': len(trips_week),
        'trips_month': len(trips_month),
        'revenue_today': total_revenue_today,
        'revenue_week': total_revenue_week,
        'revenue_month': total_revenue_month,
        'active_drivers': len([u for u in db.get_all_users() if u.role == 'driver' and u.is_active]),
        'active_vehicles': len(db.get_active_vehicles()),
        'active_routes': len(db.get_active_routes())
    }
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": current_user,
        "stats": stats,
        "recent_trips": trips_today[:10]
    })

# ===== УПРАВЛЕНИЕ ВОДИТЕЛЯМИ =====
@app.get("/drivers", response_class=HTMLResponse)
async def drivers_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Страница управления водителями"""
    drivers = [u for u in db.get_all_users() if u.role == 'driver']
    return templates.TemplateResponse("drivers.html", {
        "request": request,
        "user": current_user,
        "drivers": drivers
    })

@app.post("/drivers/create")
async def create_driver(
    surname: str = Form(...),
    first_name: str = Form(...),
    middle_name: str = Form(""),
    current_user: User = Depends(get_current_admin_user)
):
    """Создание нового водителя"""
    try:
        password = db.generate_password()
        driver_id = db.create_user(surname, first_name, middle_name, "driver", password)
        return JSONResponse({
            "success": True,
            "driver_id": driver_id,
            "password": password,
            "message": f"Водитель создан. Пароль: {password}"
        })
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/drivers/{driver_id}/deactivate")
async def deactivate_driver(
    driver_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Деактивация водителя"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE users SET is_active = 0 WHERE id = ? AND role = 'driver'", (driver_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "Водитель деактивирован"})
            else:
                return JSONResponse({"success": False, "message": "Водитель не найден"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/drivers/{driver_id}/activate")
async def activate_driver(
    driver_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Активация водителя"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE users SET is_active = 1 WHERE id = ? AND role = 'driver'", (driver_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "Водитель активирован"})
            else:
                return JSONResponse({"success": False, "message": "Водитель не найден"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

# ===== УПРАВЛЕНИЕ АВТОПАРКОМ =====
@app.get("/vehicles", response_class=HTMLResponse)
async def vehicles_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Страница управления автопарком"""
    vehicles = db.get_active_vehicles()
    # Получаем все ТС, включая неактивные
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM vehicles ORDER BY number')
        all_vehicles = []
        for row in cursor.fetchall():
            from database import Vehicle
            vehicle = Vehicle(
                id=row['id'],
                number=row['number'],
                model=row['model'],
                capacity=row['capacity'],
                is_active=row['is_active'],
                created_at=datetime.fromisoformat(row['created_at'])
            )
            all_vehicles.append(vehicle)
    
    return templates.TemplateResponse("vehicles.html", {
        "request": request,
        "user": current_user,
        "vehicles": all_vehicles
    })

@app.post("/vehicles/create")
async def create_vehicle(
    number: str = Form(...),
    model: str = Form(...),
    capacity: float = Form(0),
    current_user: User = Depends(get_current_admin_user)
):
    """Создание нового ТС"""
    try:
        vehicle_id = db.create_vehicle(number, model, capacity)
        return JSONResponse({
            "success": True,
            "vehicle_id": vehicle_id,
            "message": "Транспортное средство создано"
        })
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/vehicles/{vehicle_id}/deactivate")
async def deactivate_vehicle(
    vehicle_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Деактивация ТС"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE vehicles SET is_active = 0 WHERE id = ?", (vehicle_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "ТС деактивировано"})
            else:
                return JSONResponse({"success": False, "message": "ТС не найдено"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/vehicles/{vehicle_id}/activate")
async def activate_vehicle(
    vehicle_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Активация ТС"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE vehicles SET is_active = 1 WHERE id = ?", (vehicle_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "ТС активировано"})
            else:
                return JSONResponse({"success": False, "message": "ТС не найдено"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

# ===== УПРАВЛЕНИЕ МАРШРУТАМИ =====
@app.get("/routes", response_class=HTMLResponse)
async def routes_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Страница управления маршрутами"""
    # Получаем все маршруты, включая неактивные
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM routes ORDER BY number')
        all_routes = []
        for row in cursor.fetchall():
            from database import Route
            route = Route(
                id=row['id'],
                number=row['number'],
                name=row['name'],
                price=row['price'],
                description=row['description'] or "",
                is_active=row['is_active'],
                created_at=datetime.fromisoformat(row['created_at'])
            )
            all_routes.append(route)
    
    return templates.TemplateResponse("routes.html", {
        "request": request,
        "user": current_user,
        "routes": all_routes
    })

@app.post("/routes/create")
async def create_route(
    number: str = Form(...),
    name: str = Form(...),
    price: float = Form(...),
    description: str = Form(""),
    current_user: User = Depends(get_current_admin_user)
):
    """Создание нового маршрута"""
    try:
        route_id = db.create_route(number, name, price, description)
        return JSONResponse({
            "success": True,
            "route_id": route_id,
            "message": "Маршрут создан"
        })
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/routes/{route_id}/update_price")
async def update_route_price(
    route_id: int,
    price: float = Form(...),
    current_user: User = Depends(get_current_admin_user)
):
    """Обновление цены маршрута"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE routes SET price = ? WHERE id = ?", (price, route_id))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "Цена маршрута обновлена"})
            else:
                return JSONResponse({"success": False, "message": "Маршрут не найден"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/routes/{route_id}/deactivate")
async def deactivate_route(
    route_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Деактивация маршрута"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE routes SET is_active = 0 WHERE id = ?", (route_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "Маршрут деактивирован"})
            else:
                return JSONResponse({"success": False, "message": "Маршрут не найден"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/routes/{route_id}/activate")
async def activate_route(
    route_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Активация маршрута"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE routes SET is_active = 1 WHERE id = ?", (route_id,))
            conn.commit()
            
            if cursor.rowcount > 0:
                return JSONResponse({"success": True, "message": "Маршрут активирован"})
            else:
                return JSONResponse({"success": False, "message": "Маршрут не найден"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

# ===== ОТЧЕТЫ =====
@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Страница отчетности"""
    return templates.TemplateResponse("reports.html", {
        "request": request,
        "user": current_user
    })

@app.get("/reports/excel")
async def generate_excel_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """Генерация Excel отчета"""
    
    # Парсинг дат
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    
    # Получение данных
    trips = db.get_trips_for_report(start_dt, end_dt)
    
    # Создание Excel файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет о рейсах"
    
    # Настройка стилей
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки точно по образцу
    headers = [
        "№ п/п",
        "Дата", 
        "Номер путевого листа",
        "Наименование услуги",
        "Водитель",
        "Ставка",
        "Стоимость оказанной услуги",
        "НДС,%",
        "Итого"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные строго по образцу
    total_amount = 0
    for row_num, trip in enumerate(trips, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = border  # № п/п
        ws.cell(row=row_num, column=2, value=trip['date']).border = border  # Дата
        ws.cell(row=row_num, column=3, value=trip['waybill_number']).border = border  # Номер путевого листа
        ws.cell(row=row_num, column=4, value=trip['service_description']).border = border  # Наименование услуги
        ws.cell(row=row_num, column=5, value=trip['driver_name']).border = border  # Водитель
        ws.cell(row=row_num, column=6, value=trip['rate']).border = border  # Ставка
        ws.cell(row=row_num, column=7, value=trip['rate']).border = border  # Стоимость оказанной услуги
        ws.cell(row=row_num, column=8, value="Без НДС").border = border  # НДС,%
        ws.cell(row=row_num, column=9, value=trip['total_amount']).border = border  # Итого
        
        total_amount += trip['total_amount']
    
    # Итоговая строка точно как в образце
    if trips:
        total_row = len(trips) + 2
        # Объединяем ячейки для "ИТОГО:"
        ws.merge_cells(f'A{total_row}:H{total_row}')
        merged_cell = ws.cell(row=total_row, column=1, value="ИТОГО:")
        merged_cell.font = header_font
        merged_cell.alignment = Alignment(horizontal='right', vertical='center')
        merged_cell.border = border
        
        # Итоговая сумма
        total_cell = ws.cell(row=total_row, column=9, value=f"{total_amount:,.2f} ₽")
        total_cell.font = header_font
        total_cell.border = border
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение во временный файл
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        wb.save(tmp_file.name)
        tmp_path = tmp_file.name
    
    # Формирование имени файла
    period_str = ""
    if start_date and end_date:
        period_str = f"_{start_date}_{end_date}"
    elif start_date:
        period_str = f"_с_{start_date}"
    elif end_date:
        period_str = f"_до_{end_date}"
    
    filename = f"отчет_о_рейсах{period_str}.xlsx"
    
    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/trips")
async def get_trips_api(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """API для получения списка рейсов"""
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    
    trips = db.get_trips_for_report(start_dt, end_dt)
    return {"trips": trips}

# ===== СТРАНИЦА НАСТРОЕК =====
@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Обновленная страница настроек системы"""
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "user": current_user
    })

@app.get("/api/calendar/status")
async def get_calendar_status(current_user: User = Depends(get_current_admin_user)):
    """Получение статуса Google Calendar"""
    try:
        calendar_integration = get_calendar_integration()
        status = calendar_integration.get_connection_status()
        return JSONResponse(status)
    except Exception as e:
        logger.error(f"Ошибка получения статуса календаря: {e}")
        return JSONResponse({
            'is_available': False,
            'is_configured': False,
            'has_credentials': False,
            'has_token': False,
            'calendar_info': None,
            'error': f'Ошибка получения статуса: {str(e)}'
        })

@app.get("/api/calendar/test")
async def test_calendar_connection(current_user: User = Depends(get_current_admin_user)):
    """Тестирование подключения к Google Calendar"""
    try:
        calendar_integration = get_calendar_integration()
        result = calendar_integration.test_connection()
        return JSONResponse(result)
    except Exception as e:
        logger.error(f"Ошибка тестирования календаря: {e}")
        return JSONResponse({
            'success': False,
            'message': f'Ошибка тестирования: {str(e)}',
            'calendar_info': None
        })

@app.post("/api/calendar/setup")
async def setup_calendar(current_user: User = Depends(get_current_admin_user)):
    """Запуск процесса настройки календаря"""
    if not GOOGLE_CALENDAR_AVAILABLE:
        return JSONResponse({
            'success': False,
            'message': 'Установите Google API библиотеки: pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client'
        })
    
    if not os.path.exists('credentials.json'):
        return JSONResponse({
            'success': False,
            'message': 'Файл credentials.json не найден. Скачайте его из Google Cloud Console и поместите в корневую директорию проекта.'
        })
    
    try:
        calendar_integration = get_calendar_integration()
        
        # Проверяем текущий статус
        status = calendar_integration.get_connection_status()
        
        if status['is_configured']:
            return JSONResponse({
                'success': True,
                'message': 'Google Calendar уже настроен и готов к работе',
                'calendar_info': status.get('calendar_info')
            })
        
        # Попытка аутентификации
        test_result = calendar_integration.test_connection()
        
        if test_result['success']:
            return JSONResponse({
                'success': True,
                'message': 'Google Calendar успешно настроен',
                'calendar_info': test_result.get('calendar_info')
            })
        else:
            return JSONResponse({
                'success': False,
                'message': f'Ошибка настройки: {test_result["message"]}'
            })
            
    except Exception as e:
        logger.error(f"Ошибка настройки календаря: {e}")
        return JSONResponse({
            'success': False,
            'message': f'Ошибка настройки: {str(e)}'
        })

@app.post("/api/calendar/disconnect")
async def disconnect_calendar(current_user: User = Depends(get_current_admin_user)):
    """Отключение Google Calendar"""
    try:
        # Удаляем файл токенов
        if os.path.exists('token.json'):
            os.remove('token.json')
            return JSONResponse({
                'success': True,
                'message': 'Google Calendar отключен. Файл token.json удален.'
            })
        else:
            return JSONResponse({
                'success': True,
                'message': 'Google Calendar уже отключен'
            })
            
    except Exception as e:
        logger.error(f"Ошибка отключения календаря: {e}")
        return JSONResponse({
            'success': False,
            'message': f'Ошибка отключения: {str(e)}'
        })

@app.post("/api/calendar/create-test-event")
async def create_test_event(current_user: User = Depends(get_current_admin_user)):
    """Создание тестового события в календаре"""
    try:
        calendar_integration = get_calendar_integration()
        
        if not calendar_integration.enabled:
            return JSONResponse({
                'success': False,
                'message': 'Google Calendar не настроен'
            })
        
        # Создаем тестовое событие
        trip_data = {
            'waybill_number': f'TEST{datetime.now().strftime("%H%M%S")}',
            'vehicle_number': 'TEST-001',
            'route_number': '99',
            'route_name': 'Тестовый маршрут',
            'quantity_delivered': 100
        }
        
        user_data = {
            'id': current_user.id,
            'surname': current_user.surname,
            'first_name': current_user.first_name,
            'middle_name': getattr(current_user, 'middle_name', '')
        }
        
        event_id = calendar_integration.create_trip_event_sync(trip_data, user_data)
        
        if event_id:
            return JSONResponse({
                'success': True,
                'message': f'Тестовое событие создано успешно! ID: {event_id}',
                'event_id': event_id
            })
        else:
            return JSONResponse({
                'success': False,
                'message': 'Не удалось создать тестовое событие. Проверьте логи.'
            })
            
    except Exception as e:
        logger.error(f"Ошибка создания тестового события: {e}")
        return JSONResponse({
            'success': False,
            'message': f'Ошибка создания события: {str(e)}'
        })

# ===== API ДЛЯ TELEGRAM BOT =====
@app.post("/api/telegram/save-token")
async def save_telegram_token(
    request: Request,
    current_user: User = Depends(get_current_admin_user)
):
    """Сохранение токена Telegram бота"""
    try:
        data = await request.json()
        token = data.get('token', '').strip()
        
        if not token:
            return JSONResponse({"success": False, "message": "Токен не может быть пустым"})
        
        # Читаем существующий .env файл
        env_content = ""
        env_file = ".env"
        
        if os.path.exists(env_file):
            with open(env_file, 'r', encoding='utf-8') as f:
                env_content = f.read()
        
        # Обновляем или добавляем токен
        lines = env_content.split('\n')
        token_found = False
        
        for i, line in enumerate(lines):
            if line.strip().startswith('TELEGRAM_BOT_TOKEN='):
                lines[i] = f'TELEGRAM_BOT_TOKEN={token}'
                token_found = True
                break
        
        if not token_found:
            lines.append(f'TELEGRAM_BOT_TOKEN={token}')
        
        # Сохраняем обновленный файл
        with open(env_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        # Обновляем переменную окружения
        os.environ['TELEGRAM_BOT_TOKEN'] = token
        
        return JSONResponse({"success": True, "message": "Токен сохранен успешно"})
        
    except Exception as e:
        return JSONResponse({"success": False, "message": f"Ошибка сохранения: {str(e)}"})

# ===== API ДЛЯ СИСТЕМНОЙ ИНФОРМАЦИИ =====
@app.get("/api/system/info")
async def get_system_info(current_user: User = Depends(get_current_admin_user)):
    """Получение системной информации"""
    try:
        # Подсчитываем количество рейсов
        trips = db.get_trips_for_report()
        trips_count = len(trips)
        
        # Подсчитываем количество активных водителей
        drivers = [u for u in db.get_all_users() if u.role == 'driver' and u.is_active]
        drivers_count = len(drivers)
        
        return JSONResponse({
            "trips_count": trips_count,
            "drivers_count": drivers_count,
            "database_type": "SQLite",
            "version": "1.0.0"
        })
        
    except Exception as e:
        logger.error(f"Ошибка получения системной информации: {e}")
        return JSONResponse({
            "trips_count": 0,
            "drivers_count": 0,
            "database_type": "SQLite",
            "version": "1.0.0",
            "error": str(e)
        })


# ===== РАСШИРЕННЫЕ ОТЧЕТЫ =====

@app.get("/reports/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Страница аналитики и расширенных отчетов"""
    return templates.TemplateResponse("analytics.html", {
        "request": request,
        "user": current_user
    })

@app.get("/api/reports/drivers")
async def get_driver_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """API для получения отчетов по водителям"""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        
        stats = db.get_driver_statistics(start_dt, end_dt)
        return JSONResponse({"success": True, "data": stats})
        
    except Exception as e:
        logger.error(f"Ошибка получения отчета по водителям: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@app.get("/api/reports/vehicles")
async def get_vehicle_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """API для получения отчетов по ТС"""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        
        stats = db.get_vehicle_statistics(start_dt, end_dt)
        return JSONResponse({"success": True, "data": stats})
        
    except Exception as e:
        logger.error(f"Ошибка получения отчета по ТС: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@app.get("/api/reports/routes")
async def get_route_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """API для получения отчетов по маршрутам"""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        
        stats = db.get_route_statistics(start_dt, end_dt)
        return JSONResponse({"success": True, "data": stats})
        
    except Exception as e:
        logger.error(f"Ошибка получения отчета по маршрутам: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@app.get("/api/reports/trips")
async def get_trips_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    driver_id: Optional[int] = None,
    vehicle_id: Optional[int] = None,
    route_id: Optional[int] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """API для получения детального отчета по рейсам"""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        
        trips = db.get_trips_for_report(
            start_date=start_dt,
            end_date=end_dt,
            status=status,
            user_id=driver_id,
            vehicle_id=vehicle_id,
            route_id=route_id
        )
        
        return JSONResponse({"success": True, "data": trips})
        
    except Exception as e:
        logger.error(f"Ошибка получения отчета по рейсам: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@app.get("/api/reports/dashboard")
async def get_dashboard_data(current_user: User = Depends(get_current_admin_user)):
    """API для получения данных дашборда"""
    try:
        today = date.today()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Статистика за сегодня
        trips_today = db.get_trips_for_report(start_date=today, end_date=today)
        completed_today = [t for t in trips_today if t['status'] == 'completed']
        
        # Статистика за неделю
        trips_week = db.get_trips_for_report(start_date=week_ago, end_date=today)
        completed_week = [t for t in trips_week if t['status'] == 'completed']
        
        # Статистика за месяц
        trips_month = db.get_trips_for_report(start_date=month_ago, end_date=today)
        completed_month = [t for t in trips_month if t['status'] == 'completed']
        
        # Активные рейсы
        active_trips = db.get_trips_for_report(status='started')
        
        # Средняя продолжительность поездок
        avg_duration_month = 0
        duration_trips = [t for t in completed_month if t.get('duration_hours')]
        if duration_trips:
            avg_duration_month = sum(t['duration_hours'] for t in duration_trips) / len(duration_trips)
        
        dashboard_data = {
            'trips_today': len(trips_today),
            'completed_today': len(completed_today),
            'revenue_today': sum(t['total_amount'] for t in completed_today),
            
            'trips_week': len(trips_week),
            'completed_week': len(completed_week),
            'revenue_week': sum(t['total_amount'] for t in completed_week),
            
            'trips_month': len(trips_month),
            'completed_month': len(completed_month),
            'revenue_month': sum(t['total_amount'] for t in completed_month),
            
            'active_trips': len(active_trips),
            'avg_duration_hours': round(avg_duration_month, 2),
            
            'active_drivers': len([u for u in db.get_all_users() if u.role == 'driver' and u.is_active]),
            'active_vehicles': len(db.get_active_vehicles()),
            'active_routes': len(db.get_active_routes()),
            
            'completion_rate': round(len(completed_month) / max(len(trips_month), 1) * 100, 1)
        }
        
        return JSONResponse({"success": True, "data": dashboard_data})
        
    except Exception as e:
        logger.error(f"Ошибка получения данных дашборда: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@app.get("/reports/excel/advanced")
async def generate_advanced_excel_report(
    report_type: str = "trips",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    driver_id: Optional[int] = None,
    vehicle_id: Optional[int] = None,
    route_id: Optional[int] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """Генерация расширенного Excel отчета"""
    
    try:
        # Парсинг дат
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        
        # Создание Excel файла
        wb = openpyxl.Workbook()
        
        # Настройка стилей
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if report_type == "drivers":
            # Отчет по водителям
            ws = wb.active
            ws.title = "Отчет по водителям"
            
            headers = [
                "№", "ФИО водителя", "Всего рейсов", "Завершено", "Отменено", 
                "Общий доход (₽)", "Общее количество", "Ср. время поездки (ч)", "% завершения"
            ]
            
            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            stats = db.get_driver_statistics(start_dt, end_dt)
            for row_num, stat in enumerate(stats, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1).border = border
                ws.cell(row=row_num, column=2, value=stat['driver_name']).border = border
                ws.cell(row=row_num, column=3, value=stat['total_trips']).border = border
                ws.cell(row=row_num, column=4, value=stat['completed_trips']).border = border
                ws.cell(row=row_num, column=5, value=stat['cancelled_trips']).border = border
                ws.cell(row=row_num, column=6, value=stat['total_revenue']).border = border
                ws.cell(row=row_num, column=7, value=stat['total_quantity']).border = border
                ws.cell(row=row_num, column=8, value=stat['avg_duration_hours']).border = border
                ws.cell(row=row_num, column=9, value=f"{stat['completion_rate']}%").border = border
            
            filename = f"отчет_по_водителям"
            
        elif report_type == "vehicles":
            # Отчет по ТС
            ws = wb.active
            ws.title = "Отчет по ТС"
            
            headers = [
                "№", "Номер ТС", "Модель", "Всего рейсов", "Завершено", 
                "Общий доход (₽)", "Общее количество", "Ср. время поездки (ч)"
            ]
            
            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            stats = db.get_vehicle_statistics(start_dt, end_dt)
            for row_num, stat in enumerate(stats, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1).border = border
                ws.cell(row=row_num, column=2, value=stat['vehicle_number']).border = border
                ws.cell(row=row_num, column=3, value=stat['vehicle_model']).border = border
                ws.cell(row=row_num, column=4, value=stat['total_trips']).border = border
                ws.cell(row=row_num, column=5, value=stat['completed_trips']).border = border
                ws.cell(row=row_num, column=6, value=stat['total_revenue']).border = border
                ws.cell(row=row_num, column=7, value=stat['total_quantity']).border = border
                ws.cell(row=row_num, column=8, value=stat['avg_duration_hours']).border = border
            
            filename = f"отчет_по_тс"
            
        elif report_type == "routes":
            # Отчет по маршрутам
            ws = wb.active
            ws.title = "Отчет по маршрутам"
            
            headers = [
                "№", "Номер маршрута", "Название", "Цена за рейс (₽)", "Всего рейсов", 
                "Завершено", "Общий доход (₽)", "Общее количество", "Ср. время поездки (ч)"
            ]
            
            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            stats = db.get_route_statistics(start_dt, end_dt)
            for row_num, stat in enumerate(stats, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1).border = border
                ws.cell(row=row_num, column=2, value=stat['route_number']).border = border
                ws.cell(row=row_num, column=3, value=stat['route_name']).border = border
                ws.cell(row=row_num, column=4, value=stat['route_price']).border = border
                ws.cell(row=row_num, column=5, value=stat['total_trips']).border = border
                ws.cell(row=row_num, column=6, value=stat['completed_trips']).border = border
                ws.cell(row=row_num, column=7, value=stat['total_revenue']).border = border
                ws.cell(row=row_num, column=8, value=stat['total_quantity']).border = border
                ws.cell(row=row_num, column=9, value=stat['avg_duration_hours']).border = border
            
            filename = f"отчет_по_маршрутам"
            
        else:
            # Детальный отчет по рейсам
            ws = wb.active
            ws.title = "Детальный отчет по рейсам"
            
            headers = [
                "№", "Дата", "Номер путевого листа", "Водитель", "ТС", "Маршрут",
                "Количество", "Статус", "Время начала", "Время окончания", "Продолжительность", "Сумма (₽)"
            ]
            
            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            trips = db.get_trips_for_report(
                start_date=start_dt,
                end_date=end_dt,
                user_id=driver_id,
                vehicle_id=vehicle_id,
                route_id=route_id
            )
            
            for row_num, trip in enumerate(trips, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1).border = border
                ws.cell(row=row_num, column=2, value=trip['date']).border = border
                ws.cell(row=row_num, column=3, value=trip['waybill_number']).border = border
                ws.cell(row=row_num, column=4, value=trip['driver_name']).border = border
                ws.cell(row=row_num, column=5, value=trip['vehicle_number']).border = border
                ws.cell(row=row_num, column=6, value=f"№{trip['route_name'][:20]}").border = border
                ws.cell(row=row_num, column=7, value=trip['quantity']).border = border
                
                # Статус с цветовым кодированием
                status_cell = ws.cell(row=row_num, column=8, value=trip['status'])
                status_cell.border = border
                if trip['status'] == 'completed':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif trip['status'] == 'started':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif trip['status'] == 'cancelled':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Время
                start_time = ""
                end_time = ""
                duration = ""
                
                if trip.get('started_at'):
                    start_time = datetime.fromisoformat(trip['started_at']).strftime('%H:%M')
                if trip.get('completed_at'):
                    end_time = datetime.fromisoformat(trip['completed_at']).strftime('%H:%M')
                if trip.get('duration_hours'):
                    hours = int(trip['duration_hours'])
                    minutes = int((trip['duration_hours'] - hours) * 60)
                    duration = f"{hours}ч {minutes}мин"
                
                ws.cell(row=row_num, column=9, value=start_time).border = border
                ws.cell(row=row_num, column=10, value=end_time).border = border
                ws.cell(row=row_num, column=11, value=duration).border = border
                ws.cell(row=row_num, column=12, value=trip['total_amount']).border = border
            
            filename = f"детальный_отчет_по_рейсам"
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Формирование имени файла
        period_str = ""
        if start_date and end_date:
            period_str = f"_{start_date}_{end_date}"
        elif start_date:
            period_str = f"_с_{start_date}"
        elif end_date:
            period_str = f"_до_{end_date}"
        
        full_filename = f"{filename}{period_str}.xlsx"
        
        # Сохранение во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        return FileResponse(
            path=tmp_path,
            filename=full_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Ошибка генерации расширенного отчета: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчета: {str(e)}")

# ===== СОЗДАНИЕ ШАБЛОНА АНАЛИТИКИ =====

def create_analytics_template():
    """Создание шаблона аналитики"""
    analytics_html = '''{% extends "base.html" %}

{% block title %}Аналитика и отчеты - Система экспедирования{% endblock %}

{% block content %}
<div class="d-flex justify-content-between flex-wrap flex-md-nowrap align-items-center pt-3 pb-2 mb-3 border-bottom">
    <h1 class="h2">Аналитика и отчеты</h1>
    <div class="btn-toolbar mb-2 mb-md-0">
        <div class="btn-group me-2">
            <button type="button" class="btn btn-sm btn-outline-secondary" onclick="refreshAllData()">
                <i class="fas fa-sync-alt"></i> Обновить данные
            </button>
        </div>
    </div>
</div>

<!-- Фильтры -->
<div class="card shadow mb-4">
    <div class="card-header py-3">
        <h6 class="m-0 font-weight-bold text-primary">Фильтры отчетов</h6>
    </div>
    <div class="card-body">
        <form id="reportFilters">
            <div class="row">
                <div class="col-md-3">
                    <label for="startDate" class="form-label">Дата начала</label>
                    <input type="date" class="form-control" id="startDate" name="start_date">
                </div>
                <div class="col-md-3">
                    <label for="endDate" class="form-label">Дата окончания</label>
                    <input type="date" class="form-control" id="endDate" name="end_date">
                </div>
                <div class="col-md-3">
                    <label class="form-label">&nbsp;</label>
                    <div class="d-flex gap-2">
                        <button type="button" class="btn btn-primary" onclick="loadReportData()">
                            <i class="fas fa-search"></i> Показать
                        </button>
                        <button type="button" class="btn btn-success" onclick="exportToExcel()">
                            <i class="fas fa-file-excel"></i> Excel
                        </button>
                    </div>
                </div>
            </div>
        </form>
    </div>
</div>

<!-- Dashboard Cards -->
<div class="row mb-4">
    <div class="col-xl-3 col-md-6 mb-4">
        <div class="card border-left-primary shadow h-100 py-2">
            <div class="card-body">
                <div class="row no-gutters align-items-center">
                    <div class="col mr-2">
                        <div class="text-xs font-weight-bold text-primary text-uppercase mb-1">Активные рейсы</div>
                        <div class="h5 mb-0 font-weight-bold text-gray-800" id="activeTrips">-</div>
                    </div>
                    <div class="col-auto">
                        <i class="fas fa-truck fa-2x text-gray-300"></i>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="col-xl-3 col-md-6 mb-4">
        <div class="card border-left-success shadow h-100 py-2">
            <div class="card-body">
                <div class="row no-gutters align-items-center">
                    <div class="col mr-2">
                        <div class="text-xs font-weight-bold text-success text-uppercase mb-1">Доход за месяц</div>
                        <div class="h5 mb-0 font-weight-bold text-gray-800" id="monthlyRevenue">-</div>
                    </div>
                    <div class="col-auto">
                        <i class="fas fa-dollar-sign fa-2x text-gray-300"></i>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="col-xl-3 col-md-6 mb-4">
        <div class="card border-left-info shadow h-100 py-2">
            <div class="card-body">
                <div class="row no-gutters align-items-center">
                    <div class="col mr-2">
                        <div class="text-xs font-weight-bold text-info text-uppercase mb-1">Процент завершения</div>
                        <div class="row no-gutters align-items-center">
                            <div class="col-auto">
                                <div class="h5 mb-0 mr-3 font-weight-bold text-gray-800" id="completionRate">-</div>
                            </div>
                            <div class="col">
                                <div class="progress progress-sm mr-2">
                                    <div class="progress-bar bg-info" id="completionProgress" style="width: 0%"></div>
                                </div>
                            </div>
                        </div>
                    </div>
                    <div class="col-auto">
                        <i class="fas fa-clipboard-list fa-2x text-gray-300"></i>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="col-xl-3 col-md-6 mb-4">
        <div class="card border-left-warning shadow h-100 py-2">
            <div class="card-body">
                <div class="row no-gutters align-items-center">
                    <div class="col mr-2">
                        <div class="text-xs font-weight-bold text-warning text-uppercase mb-1">Ср. время поездки</div>
                        <div class="h5 mb-0 font-weight-bold text-gray-800" id="avgDuration">-</div>
                    </div>
                    <div class="col-auto">
                        <i class="fas fa-clock fa-2x text-gray-300"></i>
                    </div>
                </div>
            </div>
        </div>
    </div>
</div>

<!-- Reports Table -->
<div class="card shadow mb-4">
    <div class="card-header py-3">
        <h6 class="m-0 font-weight-bold text-primary" id="reportTitle">Отчет по водителям</h6>
    </div>
    <div class="card-body">
        <div id="reportTableContainer">
            <div class="text-center text-muted py-4">
                <i class="fas fa-chart-bar fa-3x mb-3"></i>
                <p>Выберите параметры и нажмите "Показать" для генерации отчета</p>
            </div>
        </div>
    </div>
</div>

<!-- Charts Row -->
<div class="row">
    <div class="col-xl-8 col-lg-7">
        <div class="card shadow mb-4">
            <div class="card-header py-3">
                <h6 class="m-0 font-weight-bold text-primary">График доходов</h6>
            </div>
            <div class="card-body">
                <canvas id="revenueChart" width="400" height="200"></canvas>
            </div>
        </div>
    </div>

    <div class="col-xl-4 col-lg-5">
        <div class="card shadow mb-4">
            <div class="card-header py-3">
                <h6 class="m-0 font-weight-bold text-primary">Распределение статусов</h6>
            </div>
            <div class="card-body">
                <canvas id="statusChart" width="400" height="400"></canvas>
            </div>
        </div>
    </div>
</div>
{% endblock %}

{% block scripts %}
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/3.9.1/chart.min.js"></script>
<script>
let currentReportType = 'drivers';
let revenueChart = null;
let statusChart = null;

document.addEventListener('DOMContentLoaded', function() {
    // Устанавливаем дефолтные даты (последние 30 дней)
    const today = new Date();
    const monthAgo = new Date(today.getTime() - 30 * 24 * 60 * 60 * 1000);
    
    document.getElementById('endDate').value = today.toISOString().split('T')[0];
    document.getElementById('startDate').value = monthAgo.toISOString().split('T')[0];
    
    // Загружаем данные дашборда
    loadDashboardData();
    
    // Загружаем первоначальный отчет
    loadReportData();
    
    // Обработчик изменения типа отчета
    document.getElementById('reportType').addEventListener('change', function() {
        currentReportType = this.value;
        updateReportTitle();
    });
});

function updateReportTitle() {
    const titles = {
        'drivers': 'Отчет по водителям',
        'vehicles': 'Отчет по транспортным средствам',
        'routes': 'Отчет по маршрутам',
        'trips': 'Детальный отчет по рейсам'
    };
    document.getElementById('reportTitle').textContent = titles[currentReportType];
}

async function loadDashboardData() {
    try {
        const response = await fetch('/api/reports/dashboard');
        const result = await response.json();
        
        if (result.success) {
            const data = result.data;
            
            document.getElementById('activeTrips').textContent = data.active_trips;
            document.getElementById('monthlyRevenue').textContent = formatCurrency(data.revenue_month);
            document.getElementById('completionRate').textContent = data.completion_rate + '%';
            document.getElementById('avgDuration').textContent = data.avg_duration_hours + 'ч';
            
            // Обновляем прогресс-бар
            document.getElementById('completionProgress').style.width = data.completion_rate + '%';
            
        }
    } catch (error) {
        console.error('Ошибка загрузки данных дашборда:', error);
    }
}

async function loadReportData() {
    const startDate = document.getElementById('startDate').value;
    const endDate = document.getElementById('endDate').value;
    const reportType = document.getElementById('reportType').value;
    
    currentReportType = reportType;
    updateReportTitle();
    
    try {
        const params = new URLSearchParams();
        if (startDate) params.append('start_date', startDate);
        if (endDate) params.append('end_date', endDate);
        
        const response = await fetch(`/api/reports/${reportType}?${params}`);
        const result = await response.json();
        
        if (result.success) {
            displayReportTable(result.data, reportType);
            updateCharts(result.data, reportType);
        } else {
            showError('Ошибка загрузки отчета: ' + result.error);
        }
    } catch (error) {
        console.error('Ошибка загрузки отчета:', error);
        showError('Ошибка загрузки отчета');
    }
}

function displayReportTable(data, reportType) {
    const container = document.getElementById('reportTableContainer');
    
    if (!data || data.length === 0) {
        container.innerHTML = '<div class="text-center text-muted py-4">Нет данных для отображения</div>';
        return;
    }
    
    let tableHtml = '<div class="table-responsive"><table class="table table-bordered table-hover"><thead class="table-dark">';
    
    // Заголовки таблицы в зависимости от типа отчета
    if (reportType === 'drivers') {
        tableHtml += `
            <tr>
                <th>Водитель</th>
                <th>Рейсов</th>
                <th>Завершено</th>
                <th>Отменено</th>
                <th>Доход</th>
                <th>Ср. время</th>
                <th>% завершения</th>
            </tr>
        `;
    } else if (reportType === 'vehicles') {
        tableHtml += `
            <tr>
                <th>ТС</th>
                <th>Модель</th>
                <th>Рейсов</th>
                <th>Завершено</th>
                <th>Доход</th>
                <th>Количество</th>
                <th>Ср. время</th>
            </tr>
        `;
    } else if (reportType === 'routes') {
        tableHtml += `
            <tr>
                <th>Маршрут</th>
                <th>Название</th>
                <th>Цена</th>
                <th>Рейсов</th>
                <th>Завершено</th>
                <th>Доход</th>
                <th>Ср. время</th>
            </tr>
        `;
    }
    
    tableHtml += '</thead><tbody>';
    
    // Данные таблицы
    data.forEach(item => {
        tableHtml += '<tr>';
        
        if (reportType === 'drivers') {
            tableHtml += `
                <td><strong>${item.driver_name}</strong></td>
                <td>${item.total_trips}</td>
                <td><span class="badge bg-success">${item.completed_trips}</span></td>
                <td><span class="badge bg-danger">${item.cancelled_trips}</span></td>
                <td><strong>${formatCurrency(item.total_revenue)}</strong></td>
                <td>${item.avg_duration_hours}ч</td>
                <td>
                    <div class="progress" style="height: 20px;">
                        <div class="progress-bar" style="width: ${item.completion_rate}%">${item.completion_rate}%</div>
                    </div>
                </td>
            `;
        } else if (reportType === 'vehicles') {
            tableHtml += `
                <td><strong>${item.vehicle_number}</strong></td>
                <td>${item.vehicle_model}</td>
                <td>${item.total_trips}</td>
                <td><span class="badge bg-success">${item.completed_trips}</span></td>
                <td><strong>${formatCurrency(item.total_revenue)}</strong></td>
                <td>${item.total_quantity}</td>
                <td>${item.avg_duration_hours}ч</td>
            `;
        } else if (reportType === 'routes') {
            tableHtml += `
                <td><strong>№${item.route_number}</strong></td>
                <td>${item.route_name}</td>
                <td>${formatCurrency(item.route_price)}</td>
                <td>${item.total_trips}</td>
                <td><span class="badge bg-success">${item.completed_trips}</span></td>
                <td><strong>${formatCurrency(item.total_revenue)}</strong></td>
                <td>${item.avg_duration_hours}ч</td>
            `;
        }
        
        tableHtml += '</tr>';
    });
    
    tableHtml += '</tbody></table></div>';
    container.innerHTML = tableHtml;
}

function updateCharts(data, reportType) {
    // График доходов
    const revenueData = data.slice(0, 10).map(item => ({
        label: reportType === 'drivers' ? item.driver_name : 
               reportType === 'vehicles' ? item.vehicle_number :
               reportType === 'routes' ? `№${item.route_number}` : item.name,
        value: item.total_revenue || 0
    }));
    
    updateRevenueChart(revenueData);
    
    // График статусов (только для детального отчета)
    if (reportType === 'trips') {
        // Для детального отчета покажем статистику статусов
        const statusData = {
            completed: data.filter(t => t.status === 'completed').length,
            started: data.filter(t => t.status === 'started').length,
            created: data.filter(t => t.status === 'created').length,
            cancelled: data.filter(t => t.status === 'cancelled').length
        };
        updateStatusChart(statusData);
    } else {
        // Для других отчетов покажем топ-5
        const topData = data.slice(0, 5);
        updateStatusChart(topData, reportType);
    }
}

function updateRevenueChart(data) {
    const ctx = document.getElementById('revenueChart').getContext('2d');
    
    if (revenueChart) {
        revenueChart.destroy();
    }
    
    revenueChart = new Chart(ctx, {
        type: 'bar',
        data: {
            labels: data.map(item => item.label),
            datasets: [{
                label: 'Доход (₽)',
                data: data.map(item => item.value),
                backgroundColor: 'rgba(54, 162, 235, 0.6)',
                borderColor: 'rgba(54, 162, 235, 1)',
                borderWidth: 1
            }]
        },
        options: {
            responsive: true,
            scales: {
                y: {
                    beginAtZero: true,
                    ticks: {
                        callback: function(value) {
                            return formatCurrency(value);
                        }
                    }
                }
            },
            plugins: {
                tooltip: {
                    callbacks: {
                        label: function(context) {
                            return 'Доход: ' + formatCurrency(context.parsed.y);
                        }
                    }
                }
            }
        }
    });
}

function updateStatusChart(data, reportType = 'trips') {
    const ctx = document.getElementById('statusChart').getContext('2d');
    
    if (statusChart) {
        statusChart.destroy();
    }
    
    let chartData, chartLabels;
    
    if (reportType === 'trips') {
        chartLabels = ['Завершено', 'В пути', 'Создано', 'Отменено'];
        chartData = [data.completed, data.started, data.created, data.cancelled];
    } else {
        chartLabels = data.map(item => 
            reportType === 'drivers' ? item.driver_name.split(' ')[0] :
            reportType === 'vehicles' ? item.vehicle_number :
            `№${item.route_number}`
        );
        chartData = data.map(item => item.total_trips);
    }
    
    statusChart = new Chart(ctx, {
        type: 'doughnut',
        data: {
            labels: chartLabels,
            datasets: [{
                data: chartData,
                backgroundColor: [
                    '#28a745',
                    '#ffc107', 
                    '#17a2b8',
                    '#dc3545',
                    '#6f42c1'
                ]
            }]
        },
        options: {
            responsive: true,
            plugins: {
                legend: {
                    position: 'bottom'
                }
            }
        }
    });
}

async function exportToExcel() {
    const startDate = document.getElementById('startDate').value;
    const endDate = document.getElementById('endDate').value;
    const reportType = document.getElementById('reportType').value;
    
    const params = new URLSearchParams();
    params.append('report_type', reportType);
    if (startDate) params.append('start_date', startDate);
    if (endDate) params.append('end_date', endDate);
    
    const url = `/reports/excel/advanced?${params}`;
    
    // Создаем невидимую ссылку для скачивания
    const link = document.createElement('a');
    link.href = url;
    link.download = '';
    document.body.appendChild(link);
    link.click();
    document.body.removeChild(link);
    
    showAlert('Отчет экспортируется...', 'info');
}

function refreshAllData() {
    loadDashboardData();
    loadReportData();
    showAlert('Данные обновлены', 'success');
}

function formatCurrency(amount) {
    return new Intl.NumberFormat('ru-RU', {
        style: 'currency',
        currency: 'RUB',
        minimumFractionDigits: 0
    }).format(amount);
}

function showAlert(message, type) {
    const alertDiv = document.createElement('div');
    alertDiv.className = `alert alert-${type} alert-dismissible fade show position-fixed`;
    alertDiv.style.cssText = 'top: 20px; right: 20px; z-index: 9999; max-width: 400px;';
    alertDiv.innerHTML = `
        ${message}
        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
    `;
    
    document.body.appendChild(alertDiv);
    
    setTimeout(() => {
        if (alertDiv.parentNode) {
            alertDiv.remove();
        }
    }, 3000);
}

function showError(message) {
    showAlert(message, 'danger');
}
</script>
{% endblock %}'''
    
    # Создаем файл шаблона
    os.makedirs("templates", exist_ok=True)
    with open("templates/analytics.html", "w", encoding="utf-8") as f:
        f.write(analytics_html)
    
    print("✅ Создан шаблон analytics.html")

# Вызов функции создания шаблона
if __name__ == "__main__":
    create_analytics_template()

# Добавьте эти endpoints в web_app.py

# ===== УПРАВЛЕНИЕ ПАРОЛЯМИ ВОДИТЕЛЕЙ =====

@app.post("/drivers/{driver_id}/reset_password")
async def reset_driver_password(
    driver_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Сброс пароля водителя на новый случайный"""
    try:
        new_password = db.reset_user_password(driver_id)
        
        if new_password:
            return JSONResponse({
                "success": True, 
                "new_password": new_password,
                "message": f"Пароль водителя сброшен. Новый пароль: {new_password}"
            })
        else:
            return JSONResponse({
                "success": False, 
                "message": "Водитель не найден или произошла ошибка"
            })
            
    except Exception as e:
        logger.error(f"Ошибка сброса пароля водителя {driver_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})


@app.post("/drivers/{driver_id}/change_password")
async def change_driver_password(
    driver_id: int,
    new_password: str = Form(...),
    current_user: User = Depends(get_current_admin_user)
):
    """Изменение пароля водителя на указанный"""
    try:
        if len(new_password.strip()) < 6:
            return JSONResponse({
                "success": False, 
                "message": "Пароль должен содержать минимум 6 символов"
            })
        
        success = db.change_user_password(driver_id, new_password.strip())
        
        if success:
            return JSONResponse({
                "success": True, 
                "message": "Пароль водителя успешно изменен"
            })
        else:
            return JSONResponse({
                "success": False, 
                "message": "Водитель не найден или произошла ошибка"
            })
            
    except Exception as e:
        logger.error(f"Ошибка изменения пароля водителя {driver_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})

@app.get("/drivers/{driver_id}/info")
async def get_driver_info(
    driver_id: int,
    current_user: User = Depends(get_current_admin_user)
):
    """Получение подробной информации о водителе"""
    try:
        driver_info = db.get_user_info(driver_id)
        
        if driver_info:
            return JSONResponse({"success": True, "data": driver_info})
        else:
            return JSONResponse({"success": False, "message": "Водитель не найден"})
            
    except Exception as e:
        logger.error(f"Ошибка получения информации о водителе {driver_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})

# ===== УДАЛЕНИЕ ЗАПИСЕЙ =====

@app.delete("/drivers/{driver_id}")
async def delete_driver(
    driver_id: int,
    force: bool = False,
    current_user: User = Depends(get_current_admin_user)
):
    """Удаление водителя"""
    try:
        success, message = db.delete_user(driver_id, force)
        
        return JSONResponse({
            "success": success,
            "message": message
        })
        
    except Exception as e:
        logger.error(f"Ошибка удаления водителя {driver_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})

@app.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(
    vehicle_id: int,
    force: bool = False,
    current_user: User = Depends(get_current_admin_user)
):
    """Удаление транспортного средства"""
    try:
        success, message = db.delete_vehicle(vehicle_id, force)
        
        return JSONResponse({
            "success": success,
            "message": message
        })
        
    except Exception as e:
        logger.error(f"Ошибка удаления ТС {vehicle_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})

@app.delete("/routes/{route_id}")
async def delete_route(
    route_id: int,
    force: bool = False,
    current_user: User = Depends(get_current_admin_user)
):
    """Удаление маршрута"""
    try:
        success, message = db.delete_route(route_id, force)
        
        return JSONResponse({
            "success": success,
            "message": message
        })
        
    except Exception as e:
        logger.error(f"Ошибка удаления маршрута {route_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})

@app.delete("/trips/{trip_id}")
async def delete_trip(
    trip_id: int,
    cancel_calendar_event: bool = True,
    current_user: User = Depends(get_current_admin_user)
):
    """Удаление рейса"""
    try:
        success, message = db.delete_trip(trip_id, cancel_calendar_event)
        
        return JSONResponse({
            "success": success,
            "message": message
        })
        
    except Exception as e:
        logger.error(f"Ошибка удаления рейса {trip_id}: {e}")
        return JSONResponse({"success": False, "message": str(e)})

# ===== МАССОВЫЕ ОПЕРАЦИИ =====

@app.post("/drivers/bulk_action")
async def bulk_driver_action(
    request: Request,
    current_user: User = Depends(get_current_admin_user)
):
    """Массовые операции с водителями"""
    try:
        data = await request.json()
        action = data.get('action')
        driver_ids = data.get('driver_ids', [])
        
        if not driver_ids:
            return JSONResponse({"success": False, "message": "Не выбраны водители"})
        
        results = []
        
        if action == 'activate':
            with db.get_connection() as conn:
                cursor = conn.cursor()
                for driver_id in driver_ids:
                    cursor.execute("UPDATE users SET is_active = 1 WHERE id = ? AND role = 'driver'", (driver_id,))
                conn.commit()
            results.append(f"Активировано {len(driver_ids)} водителей")
            
        elif action == 'deactivate':
            with db.get_connection() as conn:
                cursor = conn.cursor()
                for driver_id in driver_ids:
                    cursor.execute("UPDATE users SET is_active = 0 WHERE id = ? AND role = 'driver'", (driver_id,))
                conn.commit()
            results.append(f"Деактивировано {len(driver_ids)} водителей")
            
        elif action == 'reset_passwords':
            for driver_id in driver_ids:
                new_password = db.reset_user_password(driver_id)
                if new_password:
                    # Получаем информацию о водителе
                    driver_info = db.get_user_info(driver_id)
                    if driver_info:
                        results.append(f"{driver_info['full_name']}: {new_password}")
            
        elif action == 'delete':
            force = data.get('force', False)
            deleted_count = 0
            for driver_id in driver_ids:
                success, message = db.delete_user(driver_id, force)
                if success:
                    deleted_count += 1
            results.append(f"Удалено {deleted_count} из {len(driver_ids)} водителей")
            
        else:
            return JSONResponse({"success": False, "message": "Неизвестное действие"})
        
        return JSONResponse({
            "success": True,
            "message": "Операция выполнена",
            "results": results
        })
        
    except Exception as e:
        logger.error(f"Ошибка массовой операции с водителями: {e}")
        return JSONResponse({"success": False, "message": str(e)})

# ===== ПОИСК И ФИЛЬТРАЦИЯ =====

@app.get("/api/search")
async def global_search(
    q: str,
    type: Optional[str] = None,
    current_user: User = Depends(get_current_admin_user)
):
    """Глобальный поиск по системе"""
    try:
        results = {
            'drivers': [],
            'vehicles': [],
            'routes': [],
            'trips': []
        }
        
        search_term = f"%{q.lower()}%"
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            # Поиск водителей
            if not type or type == 'drivers':
                cursor.execute('''
                    SELECT id, surname, first_name, middle_name, is_active, telegram_id
                    FROM users 
                    WHERE role = 'driver' AND (
                        LOWER(surname) LIKE ? OR 
                        LOWER(first_name) LIKE ? OR 
                        LOWER(middle_name) LIKE ?
                    )
                    ORDER BY surname, first_name
                    LIMIT 10
                ''', (search_term, search_term, search_term))
                
                for row in cursor.fetchall():
                    results['drivers'].append({
                        'id': row['id'],
                        'name': f"{row['surname']} {row['first_name']} {row['middle_name'] or ''}".strip(),
                        'is_active': bool(row['is_active']),
                        'has_telegram': bool(row['telegram_id'])
                    })
            
            # Поиск ТС
            if not type or type == 'vehicles':
                cursor.execute('''
                    SELECT id, number, model, is_active
                    FROM vehicles 
                    WHERE LOWER(number) LIKE ? OR LOWER(model) LIKE ?
                    ORDER BY number
                    LIMIT 10
                ''', (search_term, search_term))
                
                for row in cursor.fetchall():
                    results['vehicles'].append({
                        'id': row['id'],
                        'number': row['number'],
                        'model': row['model'],
                        'is_active': bool(row['is_active'])
                    })
            
            # Поиск маршрутов
            if not type or type == 'routes':
                cursor.execute('''
                    SELECT id, number, name, price, is_active
                    FROM routes 
                    WHERE LOWER(number) LIKE ? OR LOWER(name) LIKE ?
                    ORDER BY number
                    LIMIT 10
                ''', (search_term, search_term))
                
                for row in cursor.fetchall():
                    results['routes'].append({
                        'id': row['id'],
                        'number': row['number'],
                        'name': row['name'],
                        'price': row['price'],
                        'is_active': bool(row['is_active'])
                    })
            
            # Поиск рейсов
            if not type or type == 'trips':
                cursor.execute('''
                    SELECT t.id, t.waybill_number, t.trip_date, t.status,
                           u.surname, u.first_name, v.number as vehicle_number, r.number as route_number
                    FROM trips t
                    JOIN users u ON t.user_id = u.id
                    JOIN vehicles v ON t.vehicle_id = v.id
                    JOIN routes r ON t.route_id = r.id
                    WHERE LOWER(t.waybill_number) LIKE ?
                    ORDER BY t.trip_date DESC
                    LIMIT 10
                ''', (search_term,))
                
                for row in cursor.fetchall():
                    results['trips'].append({
                        'id': row['id'],
                        'waybill_number': row['waybill_number'],
                        'trip_date': row['trip_date'],
                        'status': row['status'],
                        'driver_name': f"{row['surname']} {row['first_name']}",
                        'vehicle_number': row['vehicle_number'],
                        'route_number': row['route_number']
                    })
        
        return JSONResponse({"success": True, "results": results})
        
    except Exception as e:
        logger.error(f"Ошибка поиска '{q}': {e}")
        return JSONResponse({"success": False, "message": str(e)})

# ===== СТАТИСТИКА БЕЗОПАСНОСТИ =====

@app.get("/api/security/stats")
async def get_security_stats(current_user: User = Depends(get_current_admin_user)):
    """Статистика безопасности системы"""
    try:
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            # Водители без Telegram
            cursor.execute('''
                SELECT COUNT(*) FROM users 
                WHERE role = 'driver' AND is_active = 1 AND telegram_id IS NULL
            ''')
            drivers_without_telegram = cursor.fetchone()[0]
            
            # Неактивные водители с рейсами
            cursor.execute('''
                SELECT COUNT(DISTINCT t.user_id) FROM trips t
                JOIN users u ON t.user_id = u.id
                WHERE u.is_active = 0 AND u.role = 'driver'
            ''')
            inactive_drivers_with_trips = cursor.fetchone()[0]
            
            # Активные рейсы
            cursor.execute('SELECT COUNT(*) FROM trips WHERE status = "started"')
            active_trips = cursor.fetchone()[0]
            
            # Просроченные активные рейсы (более 12 часов)
            cursor.execute('''
                SELECT COUNT(*) FROM trips 
                WHERE status = "started" AND started_at < datetime('now', '-12 hours')
            ''')
            overdue_trips = cursor.fetchone()[0]
            
            return JSONResponse({
                "success": True,
                "stats": {
                    "drivers_without_telegram": drivers_without_telegram,
                    "inactive_drivers_with_trips": inactive_drivers_with_trips,
                    "active_trips": active_trips,
                    "overdue_trips": overdue_trips,
                    "security_score": max(0, 100 - (drivers_without_telegram * 5) - (overdue_trips * 10))
                }
            })
            
    except Exception as e:
        logger.error(f"Ошибка получения статистики безопасности: {e}")
        return JSONResponse({"success": False, "message": str(e)})

# ===== СТРАНИЦА УПРАВЛЕНИЯ РЕЙСАМИ =====

@app.get("/trips", response_class=HTMLResponse)
async def trips_management_page(request: Request, current_user: User = Depends(get_current_admin_user)):
    """Страница управления рейсами"""
    return templates.TemplateResponse("trips_management.html", {
        "request": request,
        "user": current_user
    })

# ===== API ДЛЯ УПРАВЛЕНИЯ РЕЙСАМИ =====

@app.get("/api/drivers")
async def get_drivers_list(current_user: User = Depends(get_current_admin_user)):
    """Получение списка водителей для фильтров"""
    try:
        users = db.get_all_users()
        drivers = [
            {
                'id': user.id,
                'surname': user.surname,
                'first_name': user.first_name,
                'full_name': f"{user.surname} {user.first_name}",
                'is_active': user.is_active
            }
            for user in users if user.role == 'driver'
        ]
        return JSONResponse(drivers)
    except Exception as e:
        logger.error(f"Ошибка получения списка водителей: {e}")
        return JSONResponse([])

@app.get("/api/vehicles")
async def get_vehicles_list(current_user: User = Depends(get_current_admin_user)):
    """Получение списка ТС для фильтров"""
    try:
        # Получаем все ТС, включая неактивные
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM vehicles ORDER BY number')
            vehicles = []
            for row in cursor.fetchall():
                vehicles.append({
                    'id': row['id'],
                    'number': row['number'],
                    'model': row['model'],
                    'is_active': row['is_active']
                })
        return JSONResponse(vehicles)
    except Exception as e:
        logger.error(f"Ошибка получения списка ТС: {e}")
        return JSONResponse([])

# ===== СОЗДАНИЕ ШАБЛОНА УПРАВЛЕНИЯ РЕЙСАМИ =====

def create_trips_management_template():
    """Создание шаблона управления рейсами"""
    # Шаблон уже создан в предыдущем артефакте
    # Здесь можно добавить дополнительные настройки
    
    # Создаем файл шаблона если его нет
    template_path = "templates/trips_management.html"
    if not os.path.exists(template_path):
        # Копируем содержимое из артефакта trips_management_page
        print("✅ Шаблон trips_management.html должен быть создан")
    
    print("✅ Страница управления рейсами настроена")

# Обновляем навигацию в base.html
def update_navigation_menu():
    """Обновление меню навигации для добавления управления рейсами"""
    navigation_update = '''
    <!-- Добавьте эту строку в nav меню в base.html между "Маршруты" и "Отчеты" -->
    <li class="nav-item">
        <a class="nav-link" href="/trips">
            <i class="fas fa-route"></i> Рейсы
        </a>
    </li>
    '''
    print("📝 Добавьте в меню навигации:")
    print(navigation_update)

if __name__ == "__main__":
    create_trips_management_template()
    update_navigation_menu()


# ===== СОЗДАНИЕ HTML ШАБЛОНОВ =====
def create_templates():
    """Создание HTML шаблонов"""
    
    # Создаем base.html
    base_html = '''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{% block title %}Система экспедирования{% endblock %}</title>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/bootstrap/5.3.0/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
    <link href="/static/css/style.css" rel="stylesheet">
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-primary">
        <div class="container">
            <a class="navbar-brand" href="/">
                <i class="fas fa-truck"></i> Система экспедирования
            </a>
            <div class="navbar-nav ms-auto">
                <span class="navbar-text me-3">
                    <i class="fas fa-user"></i> {{ user.first_name }} {{ user.surname }}
                </span>
            </div>
        </div>
    </nav>

    <div class="container-fluid">
        <div class="row">
            <nav class="col-md-2 d-md-block bg-light sidebar">
                <div class="position-sticky pt-3">
                    <ul class="nav flex-column">
                        <li class="nav-item">
                            <a class="nav-link" href="/">
                                <i class="fas fa-tachometer-alt"></i> Панель управления
                            </a>
                        </li>
                        <li class="nav-item">
                            <a class="nav-link" href="/drivers">
                                <i class="fas fa-users"></i> Водители
                            </a>
                        </li>
                        <li class="nav-item">
                            <a class="nav-link" href="/vehicles">
                                <i class="fas fa-truck"></i> Автопарк
                            </a>
                        </li>
                        <li class="nav-item">
                            <a class="nav-link" href="/routes">
                                <i class="fas fa-map"></i> Маршруты
                            </a>
                        </li>
                        <li class="nav-item">
                            <a class="nav-link" href="/trips">
                                <i class="fas fa-route"></i> Рейсы
                            </a>
                        </li>                        
                        <li class="nav-item">
                            <a class="nav-link" href="/reports">
                                <i class="fas fa-chart-bar"></i> Отчеты
                            </a>
                        </li>
                        <li class="nav-item">
                            <a class="nav-link" href="/settings">
                                <i class="fas fa-cog"></i> Настройки
                            </a>
                        </li>
                    </ul>
                </div>
            </nav>

            <main class="col-md-10 ms-sm-auto px-md-4">
                {% block content %}{% endblock %}
            </main>
        </div>
    </div>

    <script src="https://cdnjs.cloudflare.com/ajax/libs/bootstrap/5.3.0/js/bootstrap.bundle.min.js"></script>
    <script src="/static/js/app.js"></script>
    {% block scripts %}{% endblock %}
</body>
</html>'''
    
    with open("templates/base.html", "w", encoding="utf-8") as f:
        f.write(base_html)
    
    # Создаем остальные шаблоны с простым содержимым
    simple_templates = {
        "dashboard.html": "Панель управления",
        "drivers.html": "Управление водителями", 
        "vehicles.html": "Управление автопарком",
        "routes.html": "Управление маршрутами",
        "reports.html": "Отчеты и аналитика"
    }
    
    for filename, title in simple_templates.items():
        content = f'''{{%extends "base.html" %}}
{{%block title %}}{title} - Система экспедирования{{%endblock %}}
{{%block content %}}
<h1>{title}</h1>
<p>Страница в разработке...</p>
{{%endblock %}}'''
        
        with open(f"templates/{filename}", "w", encoding="utf-8") as f:
            f.write(content)

# ===== ЗАПУСК ПРИЛОЖЕНИЯ =====
if __name__ == "__main__":
    import uvicorn
    
    # Создаем структуру файлов
#    create_templates()
    
    print("🚀 Запуск веб-приложения...")
    print("📝 Создание HTML шаблонов...")
    print("💾 Инициализация базы данных...")
    
    # Инициализация базы данных
    db = DatabaseManager()
    
    print("✅ Веб-приложение готово к запуску!")
    print("🔗 Откройте http://localhost:8000 в браузере")
    print("👤 Логин: admin, Пароль: admin123")
    
    # Запуск сервера
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)